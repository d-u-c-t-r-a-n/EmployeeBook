'''
Created on Tuesday October 22nd 2019, 8:40:30 pm
Author: Duc Tran (duc.tran3@carleton.ca)
Carleton ID: 101158742
'''

'''
This program provides helper tools for the main GUI
'''

import pandas as pd
from openpyxl import Workbook, load_workbook
import ast

def return_next_empty_cell(file_path):
    '''
    This function return the next empty cell on column A\n
    Parameter: \n
        filepath: example (r'/Users/ductran/Documents/GUI/exportdf.xlsx')\n
    Output: Next empty cell
    '''
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    #extract last cell location
    for row in sheet.iter_rows(min_row=1,min_col=1,max_col=1):
        for cell in row:
            last_cell = cell.coordinate    
    #determine the next cell coordinate
    if len(last_cell)==2:
        if (last_cell[1]!='9'):
            next_cell = str(last_cell[0]) + str(int(last_cell[1])+1)
        else:
            next_cell = str(last_cell[0]) + str('10')
    else:
        if (last_cell[2]!='9'):
            next_cell = str(last_cell[0]) + str(last_cell[1]) + str(int(last_cell[2])+1)
        else:
            next_cell = str(last_cell[0]) + str(int(last_cell[1])+1) + str('0')
    return next_cell

def import_excel_return_dict(file_path):
    '''
    This function assume data from excel file as follows, output main keys will be date:\n
        Date        Name1   Name2   Name3   Name4\n
        20191111    Srvc3   Srvc4   Srvc1   Srvc2\n
        20191112    Srvc1   Srvc2   Srvc3   Srvc2\n

    Output: Dictionary type\n
        {20191111:\n
        {'Name1': ['Srvc3'], 'Name2': ['Srvc4], 'Name3': ['Srvc1'], 'Name4': ['Srvc3']},\n 
        20191112: \n
        {'Name1': ['Srvc1'], 'Name2': ['Srvc2], 'Name3': ['Srvc3'], 'Name4': ['Srvc2']}\n
        }
    '''
    
    xls = pd.read_excel(file_path,index_col=0)
    df = pd.DataFrame(xls).transpose()
    df_dict = df.to_dict()
    print (df_dict)
    return df_dict

def add_new_service_to_df(df,date_str,employee,service):
    '''
    This function attempts to add services to the dataframe\n
    Parameter:\n
        df: current dict of df, with main keys are date
        date_str: the string of today date
        employee: string of employee name
        service: string of service name
    Output:\n
        Return a brand new df, with new value added to the old df
    '''
    #Check if the date is new
    new_df = df
    try:
        a = (type(new_df[date_str]))
    except:
        new_df[date_str] = dict()
    print (new_df)
    #add a new dict for today data
    if (type(new_df[date_str])!=dict):        
        new_df[date_str] = dict()
    
    #add/choose an employee
    try:
        b = type(new_df[date_str][employee])
    except:
        new_df[date_str][employee] = list()
    #checking employee
    #checking the service list of the employee
    if (type(new_df[date_str][employee])==float):
        #float means empty cell
        new_df[date_str][employee] = list()
    elif (type(new_df[date_str][employee])==str):
        #str means there is already a list but in str type
        temp_list = new_df[date_str][employee]
        temp_list = ast.literal_eval(temp_list) #conver str(list) to list()
        new_df[date_str][employee] = temp_list
    
    new_df[date_str][employee].append(service)

    return new_df

def add_data_to_excel (data,file_path):
    '''
    This function parameter needs declaration\n
    This function will add more value to existing excel file,\n
    without erase the existing data
    '''    
    from openpyxl import load_workbook
    df = pd.DataFrame(data).transpose()
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(path=file_path,engine='openpyxl')
    writer.book = book

    #magic line of code
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer,index_label="Date")
    writer.save()

    
